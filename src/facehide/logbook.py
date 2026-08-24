from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path


@dataclass(frozen=True)
class LogRecord:
    when: datetime
    text: str
    person: str | None = None


def format_log_line(when: datetime, text: str) -> str:
    return f"{when.strftime('%Y-%m-%d %H:%M:%S')}  {text}"


def write_xlsx(
    path: Path,
    records: list[LogRecord],
    *,
    headers: tuple[str, str, str],
    sheet: str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    page = book.active
    page.title = sheet[:31] or "log"
    page.append(list(headers))
    for cell in page[1]:
        cell.font = Font(bold=True)
    for record in records:
        page.append([record.when, record.person or "", record.text])
    for row in page.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"
            cell.alignment = Alignment(horizontal="left")
    widths = (22, 18, 64)
    for index, width in enumerate(widths, start=1):
        page.column_dimensions[get_column_letter(index)].width = width
    page.auto_filter.ref = page.dimensions
    page.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
